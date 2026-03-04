import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "借地権評価額"

# Styles
header_fill = PatternFill("solid", fgColor="1F497D")
header_font = Font(bold=True, color="FFFFFF", size=11)
section_fill = PatternFill("solid", fgColor="DCE6F1")
section_font = Font(bold=True, color="1F497D", size=11)
result_fill = PatternFill("solid", fgColor="FFC000")
result_font = Font(bold=True, size=12)
label_font = Font(size=10)
value_font = Font(size=10)
thin = Side(style="thin", color="AAAAAA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
money_fmt = '#,##0"円"'
pct_fmt = '0%'

def cell_style(ws, row, col, value, font=None, fill=None, align="left",
               num_format=None, border=border):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center")
    if num_format:
        c.number_format = num_format
    c.border = border
    return c

# Column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 14
ws.column_dimensions["E"].width = 18

# ── Title ──────────────────────────────────────────────
ws.merge_cells("A1:E1")
c = ws["A1"]
c.value = "借地権評価額計算書（純資産価額方式）"
c.font = Font(bold=True, size=14, color="1F497D")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# ── 物件情報 ───────────────────────────────────────────
ws.merge_cells("A3:E3")
c = ws["A3"]
c.value = "【対象物件情報】"
c.font = section_font
c.fill = section_fill
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[3].height = 20

headers_info = ["No.", "項目", "内容", "", ""]
for col, h in enumerate(headers_info, 1):
    cell_style(ws, 4, col, h, font=Font(bold=True, size=10),
               fill=PatternFill("solid", fgColor="BDD7EE"), align="center")

info_rows = [
    (1, "所在",       "大畑字笠師前885番2"),
    (2, "地目",       "宅地"),
    (3, "地積",       "9,938.57㎡"),
    (4, "固定資産税評価額", "106,044,541円"),
    (5, "賃貸人",      "小松﨑雅雄"),
    (6, "賃借人",      "㈲小松崎倉庫"),
    (7, "土地の種別",   "貸家建付け地"),
]
for i, (no, label, val) in enumerate(info_rows, 5):
    ws.row_dimensions[i].height = 18
    cell_style(ws, i, 1, no,    font=label_font, align="center")
    cell_style(ws, i, 2, label, font=label_font)
    ws.merge_cells(f"C{i}:E{i}")
    cell_style(ws, i, 3, val,   font=label_font)

# ── 評価条件 ───────────────────────────────────────────
r = 13
ws.merge_cells(f"A{r}:E{r}")
c = ws[f"A{r}"]
c.value = "【評価条件】"
c.font = section_font
c.fill = section_fill
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 20

r += 1
for col, h in enumerate(["No.", "条件項目", "割合", "", ""], 1):
    cell_style(ws, r, col, h, font=Font(bold=True, size=10),
               fill=PatternFill("solid", fgColor="BDD7EE"), align="center")

cond_rows = [
    (1, "評価方式",     "倍率方式"),
    (2, "宅地評価倍率",  "1.1倍"),
    (3, "借地権割合",    "30%"),
    (4, "借家権割合",    "30%"),
    (5, "賃貸割合",     "100%"),
]
for i, (no, label, val) in enumerate(cond_rows, r + 1):
    ws.row_dimensions[i].height = 18
    cell_style(ws, i, 1, no,    font=label_font, align="center")
    cell_style(ws, i, 2, label, font=label_font)
    ws.merge_cells(f"C{i}:E{i}")
    cell_style(ws, i, 3, val,   font=label_font, align="center")

# ── 計算過程 ───────────────────────────────────────────
r = 21
ws.merge_cells(f"A{r}:E{r}")
c = ws[f"A{r}"]
c.value = "【計算過程】"
c.font = section_font
c.fill = section_fill
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 20

r += 1
for col, h in enumerate(["Step", "計算項目", "計算式", "金額", "単位"], 1):
    cell_style(ws, r, col, h, font=Font(bold=True, size=10),
               fill=PatternFill("solid", fgColor="BDD7EE"), align="center")

# Input values
fka  = 106_044_541   # 固定資産税評価額
bai  = 1.1           # 倍率
chk  = 0.30          # 借地権割合
chkk = 0.30          # 借家権割合
chin = 1.00          # 賃貸割合

jiyochi      = int(fka * bai)                             # 116,648,995
shakuchiken  = int(jiyochi * chk)                         # 34,994,698 (floor)
貸家建付け借地権 = int(shakuchiken * (1 - chkk * chin))    # 24,496,288

calc_rows = [
    (1, "自用地評価額",
     f"{fka:,}円 × {bai}",
     jiyochi, "円"),
    (2, "借地権価額",
     f"{jiyochi:,}円 × {int(chk*100)}%",
     shakuchiken, "円"),
    (3, "貸家建付け借地権評価額\n（1－借家権割合×賃貸割合）",
     f"{shakuchiken:,}円 × (1 - {int(chkk*100)}% × {int(chin*100)}%)\n＝ {shakuchiken:,}円 × 70%",
     貸家建付け借地権, "円"),
]

for i, (step, item, formula, amount, unit) in enumerate(calc_rows, r + 1):
    ws.row_dimensions[i].height = 30
    cell_style(ws, i, 1, step,    font=Font(bold=True, size=10), align="center")
    c2 = ws.cell(row=i, column=2, value=item)
    c2.font = label_font
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = border
    c3 = ws.cell(row=i, column=3, value=formula)
    c3.font = Font(size=9)
    c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c3.border = border
    cell_style(ws, i, 4, amount, font=value_font, align="right",
               num_format=money_fmt)
    cell_style(ws, i, 5, unit,  font=label_font, align="center")

# ── 結論 ───────────────────────────────────────────────
r = r + len(calc_rows) + 2
ws.merge_cells(f"A{r}:E{r}")
c = ws[f"A{r}"]
c.value = "【計上額（純資産価額方式における借地権）】"
c.font = Font(bold=True, size=12, color="FFFFFF")
c.fill = PatternFill("solid", fgColor="1F497D")
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 24

r += 1
ws.row_dimensions[r].height = 28
cell_style(ws, r, 1, "計上額", font=result_font, fill=result_fill, align="center")
ws.merge_cells(f"B{r}:C{r}")
cell_style(ws, r, 2, "貸家建付け借地権評価額", font=result_font, fill=result_fill)
cell_style(ws, r, 4, 貸家建付け借地権, font=result_font, fill=result_fill,
           align="right", num_format=money_fmt)
cell_style(ws, r, 5, "円", font=result_font, fill=result_fill, align="center")

# ── 根拠規定 ───────────────────────────────────────────
r += 2
ws.merge_cells(f"A{r}:E{r}")
c = ws[f"A{r}"]
c.value = "【根拠規定】"
c.font = section_font
c.fill = section_fill
c.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[r].height = 18

bases = [
    "・倍率方式による評価：財産評価基本通達 21の2",
    "・貸家建付け借地権の評価：財産評価基本通達 26（借地権価額×(1－借家権割合×賃貸割合)）",
    "・純資産価額方式における借地権計上：法人税基本通達 9-1-14、相続税法基本通達 186",
]
for b in bases:
    r += 1
    ws.merge_cells(f"A{r}:E{r}")
    c = ws[f"A{r}"]
    c.value = b
    c.font = Font(size=9, color="444444")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 16

output = "/home/user/test-project/借地権評価額計算書.xlsx"
wb.save(output)
print(f"Saved: {output}")
print(f"貸家建付け借地権評価額: {貸家建付け借地権:,}円")
