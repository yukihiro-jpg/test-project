"""Excel出力（色分け・罫線・印刷設定対応）"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from yokin_chosa.models import VerificationResult

# 検証結果ごとの背景色
RESULT_COLORS = {
    VerificationResult.FUND_TRANSFER.value: "C6EFCE",       # 緑
    VerificationResult.NEEDS_CONFIRMATION.value: "FFC7CE",   # 赤
    VerificationResult.CONFIRMED.value: "BDD7EE",            # 青
    VerificationResult.NOMINAL_DEPOSIT_SUSPECT.value: "FFE699",  # 橙
    VerificationResult.GIFT_TAX_ISSUE.value: "FFE699",       # 橙
}

# 共通スタイル
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
NUMBER_FORMAT = '#,##0'
DATE_FORMAT = 'yyyy/mm/dd'


def write_investigation_excel(
    sheets: dict[str, pd.DataFrame],
    output: Union[Path, BytesIO],
    movement_table_accounts_count: int = 0,
) -> None:
    """調査結果をExcelファイルに出力する

    Args:
        sheets: シート名 -> DataFrameのマッピング
        output: 出力先（ファイルパスまたはBytesIO）
        movement_table_accounts_count: 預金移動表の口座数（ヘッダ結合用）
    """
    wb = Workbook()
    # デフォルトシートを削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for sheet_name, df in sheets.items():
        if df.empty:
            continue

        ws = wb.create_sheet(title=sheet_name[:31])  # Excelのシート名は31文字制限

        # ヘッダ行
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # データ行
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row_data.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = THIN_BORDER

                # 値の設定
                if pd.isna(value) or value is None:
                    cell.value = None
                elif isinstance(value, (int, float)) and not pd.isna(value):
                    cell.value = value
                    cell.number_format = NUMBER_FORMAT
                    cell.alignment = Alignment(horizontal="right")
                else:
                    cell.value = str(value)

                # 検証結果列の色分け
                if col_name == "検証結果" and value is not None:
                    color = RESULT_COLORS.get(str(value))
                    if color:
                        cell.fill = PatternFill(
                            start_color=color, end_color=color, fill_type="solid"
                        )

        # 列幅の自動調整
        _auto_fit_columns(ws)

        # 預金移動表の場合は印刷設定
        if sheet_name == "預金移動表":
            ws.sheet_properties.pageSetUpPr = ws.sheet_properties.pageSetUpPr or None
            ws.page_setup.orientation = "landscape"
            ws.page_setup.paperSize = ws.PAPERSIZE_A3
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.print_title_rows = "1:1"  # ヘッダ行を全ページに印刷
            ws.freeze_panes = "C2"  # 日付・摘要列を固定

    if isinstance(output, BytesIO):
        wb.save(output)
    else:
        wb.save(str(output))


def _auto_fit_columns(ws) -> None:
    """列幅をコンテンツに合わせて自動調整"""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                # 日本語文字は2文字分の幅
                text = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in text)
                max_length = max(max_length, length)
        adjusted_width = min(max(max_length + 2, 8), 30)
        ws.column_dimensions[col_letter].width = adjusted_width
