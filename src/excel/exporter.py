"""Excel出力 - 罫線・色塗り付きの見やすいフォーマット."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import PropertyEvaluation

# スタイル定義
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Yu Gothic", bold=True, color="FFFFFF", size=10)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_FONT = Font(name="Yu Gothic", bold=True, size=10)
DATA_FONT = Font(name="Yu Gothic", size=10)
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
SUB_FONT = Font(name="Yu Gothic", size=9, color="444444")


def export_to_excel(evaluations: list[PropertyEvaluation]) -> bytes:
    """評価基礎情報一覧をExcelファイルとして出力.

    Returns:
        Excelファイルのバイト列
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "相続税評価 基礎情報一覧"

    # 列幅設定
    col_widths = [5, 25, 20, 10, 12, 15, 18, 10, 10, 18, 15, 15, 15, 15, 12, 12, 15, 30]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # タイトル
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
    title_cell = ws.cell(row=row, column=1, value="相続税評価 基礎情報一覧")
    title_cell.font = Font(name="Yu Gothic", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    row += 2

    # 各物件の情報を出力
    for idx, ev in enumerate(evaluations, 1):
        row = _write_property_section(ws, row, idx, ev)
        row += 1  # 物件間のスペース

    # 免責事項
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
    note_cell = ws.cell(
        row=row, column=1,
        value="※ 本情報は参考値です。正式な相続税評価には路線価図・評価明細書の確認及び税理士による検証が必要です。"
    )
    note_cell.font = Font(name="Yu Gothic", italic=True, size=9, color="666666")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def _write_property_section(
    ws, start_row: int, idx: int, ev: PropertyEvaluation
) -> int:
    """1筆分の情報をExcelに書き込み."""
    row = start_row

    # 物件番号ヘッダー
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cell = ws.cell(row=row, column=1, value=f"物件 {idx}: {ev.address}")
    cell.font = Font(name="Yu Gothic", bold=True, size=12, color="1F4E79")
    row += 1

    # --- 登記情報（謄本） ---
    row = _write_section_header(ws, row, "登記情報（謄本）")
    items = [
        ("所在", ev.location),
        ("地番", ev.chiban),
        ("登記地目", ev.chimoku_registry),
        (
            "登記地積",
            f"{ev.area_registry_sqm}㎡" if ev.area_registry_sqm is not None else "",
        ),
    ]
    for label, value in items:
        row = _write_data_row(ws, row, label, value)

    # --- 課税情報（固定資産評価証明等） ---
    row = _write_section_header(ws, row, "課税情報（固定資産評価証明等）")
    kl = ev.kotei_land
    items = [
        ("登記地目（固定資産）", kl.chimoku_registry if kl else ""),
        ("課税地目（現況）", ev.chimoku_tax),
        (
            "登記地積（固定資産）",
            f"{kl.area_registry_sqm}㎡" if kl and kl.area_registry_sqm is not None else "",
        ),
        (
            "課税地積",
            f"{ev.area_tax_sqm}㎡" if ev.area_tax_sqm is not None else "",
        ),
        (
            "固定資産税評価額",
            f"{ev.assessed_value:,}円" if ev.assessed_value is not None else "",
        ),
    ]
    for label, value in items:
        row = _write_data_row(ws, row, label, value)

    # --- 書類間整合性チェック ---
    if ev.consistency_checks:
        row = _write_section_header(ws, row, "書類間整合性チェック")
        for chk in ev.consistency_checks:
            label = f"{chk.field_name} ({chk.other_source})"
            value = chk.message
            r = _write_data_row(ws, row, label, value)
            if not chk.is_match:
                ws.cell(row=row, column=2).fill = WARN_FILL
                ws.cell(row=row, column=3).fill = WARN_FILL
            row = r

    # --- 持分情報 ---
    if ev.ownership and (ev.ownership.target_name or ev.ownership.current_share):
        row = _write_section_header(ws, row, "持分情報")
        items = [
            ("対象者", ev.ownership.target_name),
            ("基準日", ev.ownership.reference_date),
            ("基準日時点の持分", ev.ownership.current_share),
        ]
        for label, value in items:
            row = _write_data_row(ws, row, label, value)
        if ev.ownership.history_summary:
            row = _write_data_row(ws, row, "持分変遷", "")
            for entry in ev.ownership.history_summary:
                row = _write_sub_row(ws, row, entry)

    # --- 建物情報 ---
    has_building = ev.tohon_building or ev.kotei_building
    if has_building:
        row = _write_section_header(ws, row, "建物情報")
        tb = ev.tohon_building
        kb = ev.kotei_building
        kaoku = (tb.kaoku_bango if tb else "") or (kb.kaoku_bango if kb else "")
        kind = (tb.kind if tb else "") or (kb.kind if kb else "")
        structure = (tb.structure if tb else "") or (kb.structure if kb else "")
        items = [
            ("家屋番号", kaoku),
            ("種類", kind),
            ("構造", structure),
        ]
        for label, value in items:
            row = _write_data_row(ws, row, label, value)

        # 階別床面積（登記）
        if tb and tb.floor_areas:
            row = _write_data_row(ws, row, "階別床面積（登記）", "")
            for fa in tb.floor_areas:
                area_str = f"{fa.area_sqm}㎡" if fa.area_sqm is not None else ""
                row = _write_sub_row(ws, row, f"{fa.floor}: {area_str}")

        # 課税床面積
        if kb:
            row = _write_data_row(
                ws, row, "課税床面積",
                f"{kb.area_tax_sqm}㎡" if kb.area_tax_sqm is not None else "",
            )
            row = _write_data_row(
                ws, row, "建物評価額",
                f"{kb.assessed_value:,}円" if kb.assessed_value is not None else "",
            )
            row = _write_data_row(ws, row, "建築年", kb.construction_year)

    # --- 農地情報 ---
    if ev.nochi_daicho:
        nd = ev.nochi_daicho
        row = _write_section_header(ws, row, "農地情報")
        items = [
            ("農地区分", nd.farm_category),
            ("耕作者氏名", nd.farmer_name),
            ("権利種別", nd.right_type),
        ]
        for label, value in items:
            row = _write_data_row(ws, row, label, value)

    # --- 甲区（所有権）要約 ---
    ownership_history = []
    if ev.tohon_land and ev.tohon_land.ownership_history:
        ownership_history = ev.tohon_land.ownership_history
    elif ev.tohon_building and ev.tohon_building.ownership_history:
        ownership_history = ev.tohon_building.ownership_history

    if ownership_history:
        row = _write_section_header(ws, row, "甲区（所有権）要約")
        for entry in ownership_history:
            line = f"{entry.registration_date} | {entry.cause} | {entry.owner_name} | {entry.share}"
            row = _write_sub_row(ws, row, line)

    # --- 乙区（所有権以外）要約 ---
    other_rights = []
    if ev.tohon_land and ev.tohon_land.other_rights:
        other_rights = ev.tohon_land.other_rights
    elif ev.tohon_building and ev.tohon_building.other_rights:
        other_rights = ev.tohon_building.other_rights

    if other_rights:
        row = _write_section_header(ws, row, "乙区（所有権以外）要約")
        for entry in other_rights:
            line = f"{entry.registration_date} | {entry.right_type} | {entry.holder} | {entry.details}"
            row = _write_sub_row(ws, row, line)

    # --- 用途地域・都市計画 ---
    row = _write_section_header(ws, row, "用途地域・都市計画情報")
    items = [
        ("用途地域", ev.zoning.zone_type),
        ("建ぺい率", f"{ev.zoning.building_coverage_ratio}%" if ev.zoning.building_coverage_ratio else ""),
        ("容積率", f"{ev.zoning.floor_area_ratio}%" if ev.zoning.floor_area_ratio else ""),
        ("都市計画区域", ev.zoning.urban_planning_area),
    ]
    for label, value in items:
        row = _write_data_row(ws, row, label, value)

    # --- 前面道路情報 ---
    row = _write_section_header(ws, row, "前面道路情報")
    items = [
        ("道路幅員", f"{ev.road.road_width_m}m" if ev.road.road_width_m else ""),
        ("道路方位", ev.road.road_direction),
        ("道路種類", ev.road.road_type),
    ]
    for label, value in items:
        row = _write_data_row(ws, row, label, value)

    # --- ハザード情報 ---
    row = _write_section_header(ws, row, "ハザード情報")
    hazard_items = [
        ("洪水浸水想定", ev.hazard.flood_risk),
        ("土砂災害警戒", ev.hazard.landslide_risk),
        ("津波浸水想定", ev.hazard.tsunami_risk),
        ("高潮浸水想定", ev.hazard.storm_surge_risk),
    ]
    for label, value in hazard_items:
        r = _write_data_row(ws, row, label, value or "該当なし")
        # リスクありの場合は黄色ハイライト
        if value:
            ws.cell(row=row, column=2).fill = WARN_FILL
            ws.cell(row=row, column=3).fill = WARN_FILL
        row = r

    # --- 路線価/倍率情報 ---
    row = _write_section_header(ws, row, "路線価/倍率情報（国税庁）")
    area_type = "路線価地域" if ev.multiplier.is_rosenka_area else "倍率地域"
    items = [
        ("評価方式", area_type),
        ("町名", ev.multiplier.town_name),
        ("適用地域名", ev.multiplier.area_name),
        ("借地権割合", ev.multiplier.leasehold_ratio),
        ("宅地倍率", ev.multiplier.residential_multiplier),
        ("田倍率", ev.multiplier.paddy_multiplier),
        ("畑倍率", ev.multiplier.field_multiplier),
        ("山林倍率", ev.multiplier.forest_multiplier),
        ("原野倍率", ev.multiplier.wasteland_multiplier),
    ]
    for label, value in items:
        row = _write_data_row(ws, row, label, value)

    # --- 倍率方式 相続税評価額 ---
    if ev.valuation:
        v = ev.valuation
        row = _write_section_header(ws, row, "相続税評価額（倍率方式）")
        items = [
            ("評価方式", v.method),
            ("評価地目", v.chimoku_used),
            ("適用倍率", v.multiplier_raw),
            ("固定資産税評価額", f"{v.assessed_value:,}円" if v.assessed_value is not None else ""),
            ("相続税評価額（持分前）", f"{v.evaluated_value:,}円" if v.evaluated_value is not None else ""),
            ("持分", f"{v.share_fraction:.4f}" if v.share_fraction is not None else "単独所有"),
            ("相続税評価額（持分後）", f"{v.final_value:,}円" if v.final_value is not None else ""),
            ("計算式", v.formula),
        ]
        for label, value in items:
            row = _write_data_row(ws, row, label, value)
        # 注意事項
        for w in v.warnings:
            r = _write_data_row(ws, row, "注意", w)
            ws.cell(row=row, column=2).fill = WARN_FILL
            ws.cell(row=row, column=3).fill = WARN_FILL
            row = r

    # --- データソース ---
    row = _write_section_header(ws, row, "データソース")
    for src in ev.data_sources:
        ws.cell(row=row, column=2, value=src).font = DATA_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        row += 1

    # --- 注記 ---
    if ev.notes:
        row = _write_section_header(ws, row, "注記")
        for note in ev.notes:
            ws.cell(row=row, column=2, value=note).font = Font(name="Yu Gothic", size=9, italic=True)
            ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1

    return row


def _write_section_header(ws, row: int, title: str) -> int:
    """セクションヘッダーを書き込み."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_LEFT
    for col in range(2, 4):
        ws.cell(row=row, column=col).fill = SECTION_FILL
        ws.cell(row=row, column=col).border = THIN_BORDER
    return row + 1


def _write_data_row(ws, row: int, label: str, value) -> int:
    """データ行を書き込み."""
    label_cell = ws.cell(row=row, column=2, value=label)
    label_cell.font = Font(name="Yu Gothic", bold=True, size=10)
    label_cell.border = THIN_BORDER
    label_cell.alignment = ALIGN_LEFT

    value_cell = ws.cell(row=row, column=3, value=value if value else "")
    value_cell.font = DATA_FONT
    value_cell.border = THIN_BORDER
    value_cell.alignment = ALIGN_LEFT

    return row + 1


def _write_sub_row(ws, row: int, text: str) -> int:
    """サブ行（インデント付き）を書き込み."""
    cell = ws.cell(row=row, column=3, value=text)
    cell.font = SUB_FONT
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_LEFT
    return row + 1
