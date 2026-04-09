"""倍率表Excelファイル取り込みスクリプト.

`data/multipliers/` 配下のExcelファイルを読み込み、
`data/multipliers_imported.json` に市町村ごとの倍率データを書き出す。

Excelフォーマット想定:
- 1ブック = 1都道府県（または複数市町村セット）
- 1シート = 1市町村 (シート名 = "水戸市" 等)
- 行1: タイトル (令和N年分 倍率表 （市名） XX税務署)
- 行2: ヘッダ (町名, 適用地域名, 借地権割合, 宅地, 田, 畑, 山林, 原野, 牧場, 池沼)
- 行3: 空行
- 行4+: データ
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parent.parent
MULTIPLIER_DIR = REPO_ROOT / "data" / "multipliers"
OUTPUT_FILE = REPO_ROOT / "data" / "multipliers_imported.json"

COLUMN_KEYS = [
    "town_name",      # 1: 町（丁目）又は大字名
    "area_name",      # 2: 適用地域名
    "leasehold_ratio",  # 3: 借地権割合
    "residential",    # 4: 宅地
    "paddy",          # 5: 田
    "field",          # 6: 畑
    "forest",         # 7: 山林
    "wasteland",      # 8: 原野
    "pasture",        # 9: 牧場
    "pond",           # 10: 池沼
]


def _clean(v: Any) -> str:
    """セル値を文字列に正規化."""
    if v is None:
        return ""
    s = str(v).strip()
    # セル内改行を除去
    s = s.replace("\n", "").replace("\r", "")
    # 借地権割合の「―」やスペース等を空文字扱い
    if s in ("―", "-", "—"):
        return ""
    return s


def _is_header_or_title(row: tuple) -> bool:
    """ヘッダ行かタイトル行かを判定."""
    joined = "".join(_clean(v) for v in row)
    if not joined:
        return True
    if "倍率表" in joined and "令和" in joined:
        return True
    if "町" in (_clean(row[0]) if row else "") and "適用地域名" in joined:
        return True
    if _clean(row[0]).startswith("町"):
        return True
    return False


def import_sheet(ws) -> list[dict[str, str]]:
    """1シート分の倍率データを抽出."""
    records: list[dict[str, str]] = []
    last_town = ""

    for row in ws.iter_rows(min_row=1, values_only=True):
        # ヘッダ行・空行スキップ
        if _is_header_or_title(row):
            continue

        cells = [_clean(c) for c in row]
        # 10列に揃える（不足はパディング、余剰は無視）
        cells = (cells + [""] * 10)[:10]

        # 町名引き継ぎ: 空なら直前の町名を使う
        if cells[0]:
            last_town = cells[0]
        else:
            cells[0] = last_town

        # 適用地域名も町名も空ならスキップ
        if not cells[0] and not cells[1]:
            continue
        # 全カラム空ならスキップ
        if not any(cells[1:]):
            continue

        record = dict(zip(COLUMN_KEYS, cells))
        records.append(record)

    return records


def import_workbook(xlsx_path: Path) -> dict[str, list[dict[str, str]]]:
    """1ワークブック全シートを取り込み市町村名→レコード一覧の辞書を返す."""
    print(f"[読込] {xlsx_path.name}")
    wb = load_workbook(xlsx_path, data_only=True)
    result: dict[str, list[dict[str, str]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        municipality = sheet_name.strip()
        records = import_sheet(ws)
        if records:
            result[municipality] = records
            print(f"  {municipality}: {len(records)} 件")
        else:
            print(f"  {municipality}: スキップ（データなし）")

    return result


def main() -> None:
    if not MULTIPLIER_DIR.exists():
        print(f"エラー: {MULTIPLIER_DIR} が見つかりません")
        return

    xlsx_files = sorted(MULTIPLIER_DIR.glob("*.xlsx"))
    if not xlsx_files:
        print(f"{MULTIPLIER_DIR} にExcelファイルがありません")
        return

    combined: dict[str, list[dict[str, str]]] = {}
    for xlsx in xlsx_files:
        data = import_workbook(xlsx)
        for municipality, records in data.items():
            if municipality in combined:
                # 既存データがあれば上書き警告
                print(f"  ⚠ {municipality} は既に取込済み、上書きします")
            combined[municipality] = records

    # 出力
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(
        json.dumps(combined, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    total = sum(len(v) for v in combined.values())
    print(f"\n=== 完了 ===")
    print(f"出力: {OUTPUT_FILE}")
    print(f"市町村数: {len(combined)}")
    print(f"総レコード数: {total}")


if __name__ == "__main__":
    main()
